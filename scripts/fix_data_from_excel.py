#!/usr/bin/env python3
"""Fix invoice data from the Excel source file.

The import created duplicate records from Impayés columns and set
paid_amount=0 everywhere. This script:
1. Reads the Excel file
2. Removes duplicate factures (Impayés column duplicates)
3. Sets paid_amount = montant_facture - impayé_amount
4. Sets outstanding_amount = impayé_amount
5. Creates paiement records for paid amounts
"""
import asyncio
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

# Add backend to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))
from app.db.session import engine


EXCEL_PATH = Path(__file__).resolve().parent.parent / "data" / "CAMTEL_TABLEAU_SUIVI_RECOUVREMENT.xlsx"


async def main():
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active

    # Parse header row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")
    print(f"  Headers: {headers}")

    # Find facture/impayé column pairs
    months = []
    for i, h in enumerate(headers):
        if "facture" in h.lower() or "impay" in h.lower():
            # Extract month name
            parts = h.split("_")
            month_name = parts[0] if parts else h
            months.append((i, h, month_name, "facture" in h.lower()))

    print(f"  Found {len(months)} amount columns")

    # Build account -> month -> {facture, impaye} mapping
    account_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        num_compte = str(row[0]).strip()
        if num_compte not in account_data:
            account_data[num_compte] = {}
        for col_idx, col_name, month_name, is_facture in months:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or val == "":
                val = 0
            else:
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    val = 0
            if month_name not in account_data[num_compte]:
                account_data[num_compte][month_name] = {"facture": 0, "impaye": 0}
            if is_facture:
                account_data[num_compte][month_name]["facture"] = val
            else:
                account_data[num_compte][month_name]["impaye"] = val

    print(f"  Parsed {len(account_data)} accounts")

    # Now fix the database
    async with engine.begin() as conn:
        # Step 1: Count current state
        result = await conn.execute(text("SELECT COUNT(*) FROM facture"))
        current_count = result.scalar()
        print(f"\nCurrent factures: {current_count}")

        # Step 2: Delete duplicate factures (the ones from Impayés columns)
        # These have id_facture containing 'Impayés' or 'Impaye'
        result = await conn.execute(text("""
            DELETE FROM facture WHERE id_facture LIKE '%mpay%'
        """))
        deleted = result.rowcount
        print(f"Deleted {deleted} duplicate facture records (Impayés columns)")

        # Step 3: Now update paid_amount and outstanding_amount
        # based on the Excel data
        # First, get the month mapping from the facture id_facture pattern
        # id_facture format: FAC_{num_compte}_{month_name}_{year}
        result = await conn.execute(text("SELECT id_facture, num_compte, montant_facture FROM facture LIMIT 5"))
        sample = result.fetchall()
        print(f"  Sample factures: {sample[:3]}")

        # Update paid_amount = montant_facture where we know the account paid
        # (where impayé = 0 in Excel, paid = montant_facture)
        # This is complex - let's use a simpler approach:
        # For each facture, look up the Excel data and compute paid

        # First build a lookup: (num_compte, month_key) -> {facture, impaye}
        month_map = {
            "Décembre": "2025-12", "Janvier": "2026-01", "Février": "2026-02",
            "Mars": "2026-03", "Avril": "2026-04", "Mai": "2026-05", "Juin": "2026-06",
        }

        # Get all factures
        result = await conn.execute(text("SELECT id_facture, num_compte, montant_facture FROM facture"))
        all_factures = result.fetchall()
        print(f"\nProcessing {len(all_factures)} factures...")

        updates = 0
        payments_to_create = []

        for id_facture, num_compte, montant in all_factures:
            # Parse month from id_facture
            parts = id_facture.split("_")
            if len(parts) < 4:
                continue
            month_str = parts[-1]  # e.g., "Décembre2025"
            # Try to find the month name
            excel_data = account_data.get(str(num_compte).strip(), {})

            # Match month
            impaye = 0
            for month_name, data in excel_data.items():
                if month_name in month_str or month_str.startswith(month_name):
                    impaye = data.get("impaye", 0)
                    break

            paid = (montant or 0) - impaye
            if paid < 0:
                paid = 0
            outstanding = impaye

            if paid > 0:
                status = "PAID"
            else:
                status = "OPEN"

            updates += 1

            # Queue payment creation
            if paid > 0:
                payments_to_create.append({
                    "id_facture": id_facture,
                    "montant_paye": paid,
                })

        print(f"  Prepared {updates} updates, {len(payments_to_create)} payments")

        # Batch update factures - set paid_amount and outstanding_amount
        # We'll do this with raw SQL for performance
        print("  Updating factures...")

        # For a simpler approach: set paid_amount = 0 and outstanding = montant for all
        # then compute based on Excel data
        # Actually, let's just update the ones we can match

        # Create paiement records for paid amounts
        if payments_to_create:
            print(f"  Creating {len(payments_to_create)} payment records...")
            # Create in batches
            batch_size = 1000
            for i in range(0, len(payments_to_create), batch_size):
                batch = payments_to_create[i:i+batch_size]
                for p in batch:
                    # Check if payment already exists
                    result = await conn.execute(
                        text("SELECT COUNT(*) FROM paiement WHERE id_facture = :fid"),
                        {"fid": p["id_facture"]}
                    )
                    if result.scalar() == 0:
                        payment_id = f"PAY_{p['id_facture']}"
                        await conn.execute(text("""
                            INSERT INTO paiement (id_paiement, id_facture, date_paiement, montant_paye)
                            VALUES (:pid, :fid, CURRENT_DATE, :amt)
                        """), {"pid": payment_id, "fid": p["id_facture"], "amt": p["montant_paye"]})

        # Now update all factures with correct paid/outstanding
        print("  Computing paid/outstanding from paiements...")
        await conn.execute(text("""
            UPDATE facture f
            SET paid_amount = COALESCE((
                SELECT SUM(p.montant_paye) FROM paiement p WHERE p.id_facture = f.id_facture
            ), 0),
            outstanding_amount = GREATEST(f.montant_facture - COALESCE((
                SELECT SUM(p.montant_paye) FROM paiement p WHERE p.id_facture = f.id_facture
            ), 0), 0),
            status = CASE
                WHEN COALESCE((SELECT SUM(p.montant_paye) FROM paiement p WHERE p.id_facture = f.id_facture), 0) >= f.montant_facture THEN 'PAID'
                ELSE 'OPEN'
            END
        """))

        # Verify
        result = await conn.execute(text("SELECT COUNT(*) FROM facture"))
        final_count = result.scalar()
        result = await conn.execute(text("SELECT COUNT(*) FROM paiement"))
        payment_count = result.scalar()
        result = await conn.execute(text("SELECT COUNT(*) FROM facture WHERE status = 'PAID'"))
        paid_count = result.scalar()
        result = await conn.execute(text("SELECT COUNT(*) FROM facture WHERE status = 'OPEN'"))
        open_count = result.scalar()

        print(f"\n=== RESULTS ===")
        print(f"  Factures: {final_count}")
        print(f"  Paiements: {payment_count}")
        print(f"  PAID: {paid_count}")
        print(f"  OPEN: {open_count}")

        # Sample
        result = await conn.execute(text("""
            SELECT f.id_facture, f.num_compte, f.montant_facture, f.paid_amount, f.outstanding_amount, f.status
            FROM facture f ORDER BY f.num_compte, f.date_emission LIMIT 10
        """))
        print("\n  Sample:")
        for row in result:
            print(f"    {row[0]}: {row[2]} paid={row[3]} outstanding={row[4]} status={row[5]}")


if __name__ == "__main__":
    asyncio.run(main())
